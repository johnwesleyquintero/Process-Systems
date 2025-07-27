# ==============================================================================
# The Pythonic Blacksmith v4.0 (The Quartermaster's Forge)
#
# Mission: To programmatically forge the "Restock Recommender," an intelligent
#          Excel template that turns chaotic inventory data into a clear,
#          actionable restocking plan.
#
# Coded by: WesAI, Chief of Staff to the ScaleSmart Empire
# ==============================================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURATION ---
FILENAME = "Restock_Recommender_v1.0.xlsx"
MAX_ROWS = 500
DAYS_OF_STOCK_TARGET = 60 # Our 60-day goal

# --- STYLES ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16)
NOTE_FONT = Font(italic=True, color="808080")

ACTION_NEEDED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEALTHY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

def set_header_styles(sheet, max_col):
    """Applies standard header styling."""
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    sheet.freeze_panes = 'A2'

def create_instructions_sheet(wb):
    """Creates the 'Instructions' tab with dual-report instructions."""
    sheet = wb.active
    sheet.title = "Instructions"
    
    sheet['A1'] = "How to Use the Restock Recommender"
    sheet['A1'].font = TITLE_FONT

    instructions = [
        ("CRITICAL STEP 1: Download Reports", "This tool requires TWO separate reports from Seller Central."),
        ("Report A: FBA Inventory", "Go to: Reports > Fulfillment > Inventory > FBA Inventory. Click 'Request .csv Download'."),
        ("Report B: Business Report", "Go to: Reports > Business Reports > Detail Page Sales and Traffic by Child Item. Set date to 'Last 30 Days' and download."),
        ("STEP 2: Input Data", "Paste the entire contents of Report A into the 'Data_Input_Inventory' tab."),
        ("", "Paste the entire contents of Report B into the 'Data_Input_BizRpt' tab."),
        ("STEP 3: Analyze Dashboard", "Go to the 'Restock_Dashboard' tab. It will automatically calculate which products need restocking to maintain a 60-day supply.")
    ]
    
    row_num = 3
    for title, desc in instructions:
        sheet[f'A{row_num}'] = title
        sheet[f'A{row_num}'].font = Font(bold=True)
        sheet[f'B{row_num}'] = desc
        row_num += 1
    
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 100

def create_data_input_sheets(wb):
    """Creates the TWO data input tabs."""
    for name, report_name in [("Inventory", "FBA Inventory"), ("BizRpt", "Business Report")]:
        sheet = wb.create_sheet(f"Data_Input_{name}")
        sheet['A1'] = f"PASTE YOUR RAW '{report_name}' DATA EXPORT HERE"
        sheet['A1'].font = NOTE_FONT

def create_dashboard_sheet(wb):
    """Creates the main Restock Dashboard with VLOOKUP data blending."""
    sheet = wb.create_sheet("Restock_Dashboard")

    headers = [
        "ASIN", "SKU", "Item Name", "Listing Link", "Available", "Inbound", 
        "Total Stock", "Sold (Last 30d)", "Daily Sales (Avg)", f"{DAYS_OF_STOCK_TARGET}-Day Target", 
        "RECOMMENDED UNITS", "Status"
    ]
    sheet.append(headers)
    set_header_styles(sheet, len(headers))

    # NOTE: The VLOOKUP column numbers are ESTIMATES. You may need to adjust them
    # based on the actual columns in the downloaded Amazon reports.
    for i in range(2, MAX_ROWS + 2):
        # --- Data Pulled from INVENTORY Report (using SKU as key) ---
        sku_ref = f"Data_Input_Inventory!D{i}" # Assuming SKU is in column D
        sheet[f'B{i}'] = f'=IF(ISBLANK({sku_ref}),"",{sku_ref})'
        sheet[f'C{i}'] = f'=IFNA(VLOOKUP(B{i}, Data_Input_Inventory!D:F, 3, FALSE), "")' # Item Name
        sheet[f'A{i}'] = f'=IFNA(VLOOKUP(B{i}, Data_Input_Inventory!D:J, 7, FALSE), "")' # ASIN
        sheet[f'E{i}'] = f'=IFNA(VLOOKUP(B{i}, Data_Input_Inventory!D:K, 8, FALSE), 0)' # Available
        sheet[f'F{i}'] = f'=IFNA(VLOOKUP(B{i}, Data_Input_Inventory!D:L, 9, FALSE), 0)' # Inbound

        # --- Data Pulled from BUSINESS Report (using ASIN as key) ---
        sheet[f'H{i}'] = f'=IFNA(VLOOKUP(A{i}, Data_Input_BizRpt!B:R, 17, FALSE), 0)' # Units Ordered
        
        # --- The Architect's Calculations ---
        sheet[f'D{i}'] = f'=IF(A{i}<>"", HYPERLINK("https://www.amazon.com/dp/"&A{i}, "View on Amazon"), "")' # Listing Link
        sheet[f'G{i}'] = f'=E{i}+F{i}' # Total Stock (Available + Inbound)
        sheet[f'I{i}'] = f'=H{i}/30' # Daily Sales Average
        sheet[f'J{i}'] = f'=I{i}*{DAYS_OF_STOCK_TARGET}' # 60-Day Target Stock Level
        
        # The Final, Beautiful Recommendation Formula
        sheet[f'K{i}'] = f'=IF(A{i}<>"", ROUND(MAX(0, J{i}-G{i}), 0), "")' # RECOMMENDED UNITS
        
        # The Status
        sheet[f'L{i}'] = f'=IF(K{i}>0, "Restock Needed", "Healthy")'

    # Conditional Formatting to highlight what needs action
    sheet.conditional_formatting.add(f'K2:K{MAX_ROWS + 1}', CellIsRule(operator='greaterThan', formula=['0'], fill=ACTION_NEEDED_FILL))
    sheet.conditional_formatting.add(f'L2:L{MAX_ROWS + 1}', CellIsRule(operator='equal', formula=['"Healthy"'], fill=HEALTHY_FILL))

    # Adjust column widths
    for col_letter in ['A', 'B', 'C', 'D', 'K', 'L']:
        sheet.column_dimensions[col_letter].width = 25
        
def main():
    """Main function to forge the report."""
    print("Firing up the Pythonic Blacksmith (v4.0)...")
    print(f"Forging the Quartermaster's Blade: {FILENAME}")
    
    wb = openpyxl.Workbook()
    
    create_instructions_sheet(wb)
    create_data_input_sheets(wb)
    create_dashboard_sheet(wb)
    
    wb.save(FILENAME)
    print("\nForge complete. The weapon is ready for deployment.")
    print(f"'{FILENAME}' has been created.")

if __name__ == "__main__":
    main()
