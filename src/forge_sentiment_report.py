# ==============================================================================
# The Pythonic Blacksmith v2.0 (The Oracle's Blade)
#
# Mission: To programmatically forge the "Listing Health & Sentiment Report"
#          Excel template, turning a "freestyle" request into a structured
#          analytical system.
#
# Coded by: WesAI, Chief of Staff to the ScaleSmart Empire
# ==============================================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# --- CONFIGURATION ---
FILENAME = "Listing_Health_Sentiment_Report_v1.0.xlsx"

# --- STYLES (Expanded) ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16)
SUB_TITLE_FONT = Font(bold=True, size=12)
NOTE_FONT = Font(italic=True, color="808080")
BOX_BORDER = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

def apply_styles(cell, font=None, fill=None, alignment=None, border=None):
    """Utility to apply multiple styles to a cell."""
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border

def create_instructions_summary_sheet(wb):
    """Creates the main summary and instructions tab."""
    sheet = wb.active
    sheet.title = "Instructions_Summary"

    sheet['A1'] = "Listing Health & Sentiment Report"
    sheet['A1'].font = TITLE_FONT

    sheet['A3'] = "Executive Summary (Manual Entry)"
    sheet['A3'].font = SUB_TITLE_FONT
    sheet.merge_cells('A4:H10')
    apply_styles(sheet['A4'], alignment=Alignment(vertical='top', wrap_text=True), border=BOX_BORDER)
    sheet['A4'] = "After analyzing an ASIN in the 'ASIN_Deep_Dive' tab, write your final conclusions here."

    sheet['A12'] = "How to Use This Report"
    sheet['A12'].font = SUB_TITLE_FONT

    instructions = [
        ("1. Input Data", "Paste the raw data exports from Seller Central into the corresponding 'Data_Input' tabs."),
        ("2. Set Targets", "Go to the 'Config_Lists' tab and paste your list of Top 10 ASINs into column A."),
        ("3. Analyze", "Go to the 'ASIN_Deep_Dive' tab and select an ASIN from the dropdown in cell B1."),
        ("4. Conclude", "The dashboard will auto-populate. Analyze the data, and write your final recommendations in the Executive Summary section above.")
    ]
    row_num = 13
    for title, desc in instructions:
        sheet[f'A{row_num}'] = title
        sheet[f'A{row_num}'].font = Font(bold=True)
        sheet[f'B{row_num}'] = desc
        row_num += 1

    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 100

def create_data_input_sheets(wb):
    """Creates all necessary data input tabs."""
    for name in ["VOC", "BizRpt", "Comp"]:
        sheet = wb.create_sheet(f"Data_Input_{name}")
        sheet['A1'] = f"PASTE YOUR RAW '{name}' DATA EXPORT HERE"
        apply_styles(sheet['A1'], font=NOTE_FONT)

def create_config_sheet(wb):
    """Creates the hidden configuration sheet."""
    sheet = wb.create_sheet("Config_Lists")
    sheet['A1'] = "Top 10 ASINs List"
    sheet['A1'].font = Font(bold=True)
    sheet['A2'] = "(Paste your list of target ASINs below)"
    sheet['A2'].font = NOTE_FONT
    # sheet.sheet_state = 'hidden' # Uncomment to make the sheet hidden by default

def create_deep_dive_dashboard(wb):
    """Creates the main ASIN Deep Dive dashboard."""
    sheet = wb.create_sheet("ASIN_Deep_Dive")

    # --- ASIN Selector ---
    sheet['A1'] = "Select ASIN to Analyze:"
    sheet['A1'].font = SUB_TITLE_FONT

    # Create the dropdown list
    dv = DataValidation(type="list", formula1="=Config_Lists!$A$3:$A$12")
    sheet.add_data_validation(dv)
    dv.add('B1')
    sheet['B1'].value = "<- SELECT HERE"
    sheet['B1'].font = Font(bold=True, color="FF0000")

    # --- Dashboard Layout ---
    sections = {
        "A3": "Key Performance Metrics (from Business Report)",
        "E3": "Customer Sentiment (from Voice of the Customer)",
        "A10": "Competitor Snapshot (Manual Input from 'Data_Input_Comp')",
        "A15": "Sentiment Analysis Summary (Manual Entry)",
        "A22": "Final Recommendation (Manual Entry)"
    }
    for cell, title in sections.items():
        sheet[cell] = title
        sheet[cell].font = SUB_TITLE_FONT

    # --- Key Performance Metrics ---
    metrics_kpis = ["Product Name", "Sessions", "Conversion Rate", "Buy Box %", "Ordered Sales"]
    for i, kpi in enumerate(metrics_kpis, 4):
        sheet[f'A{i}'] = kpi
        # The VLOOKUP magic (adjust column numbers as needed)
        sheet[f'B{i}'] = f'=IFNA(VLOOKUP($B$1, Data_Input_BizRpt!A:Z, {i-1}, FALSE), "Not Found")'

    # --- Customer Sentiment ---
    sentiment_kpis = ["CX Health", "NCX Rate", "Top NCX Reason"]
    for i, kpi in enumerate(sentiment_kpis, 4):
        sheet[f'E{i}'] = kpi
        sheet[f'F{i}'] = f'=IFNA(VLOOKUP($B$1, Data_Input_VOC!A:Z, {i+2}, FALSE), "Not Found")'

    # --- Manual Entry Sections ---
    sheet.merge_cells('B15:H20')
    apply_styles(sheet['B15'], alignment=Alignment(vertical='top', wrap_text=True), border=BOX_BORDER)
    sheet['B15'] = "Synthesize findings from Helium 10 review analysis here..."

    sheet.merge_cells('B22:H27')
    apply_styles(sheet['B22'], alignment=Alignment(vertical='top', wrap_text=True), border=BOX_BORDER)
    sheet['B22'] = "Based on all data, the final recommended action is..."

    # Column widths for aesthetics
    for col in ['A', 'E']: sheet.column_dimensions[col].width = 30


def main():
    """Main function to forge the report."""
    print("Firing up the Pythonic Blacksmith (v2.0)...")
    print(f"Forging the Oracle's Blade: {FILENAME}")

    wb = openpyxl.Workbook()

    create_instructions_summary_sheet(wb)
    create_data_input_sheets(wb)
    create_config_sheet(wb)
    create_deep_dive_dashboard(wb)

    wb.save(FILENAME)
    print("\nForge complete. The weapon is ready.")
    print(f"'{FILENAME}' has been created in your directory.")

if __name__ == "__main__":
    main()
